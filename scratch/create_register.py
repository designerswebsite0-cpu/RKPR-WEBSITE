import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Define output path
output_path = "Website Sources Register.xlsx"

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sources Register"

# Ensure grid lines are visible
ws.views.sheetView[0].showGridLines = True

# Styling
title_font = Font(name="Calibri", size=15, bold=True, color="1E3A2F")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1E3A2F", end_color="1E3A2F", fill_type="solid")
cell_font = Font(name="Calibri", size=10)
bold_font = Font(name="Calibri", size=10, bold=True)
center_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")

thin_side = Side(border_style="thin", color="CCCCCC")
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# Title row
ws.merge_cells("A1:H1")
ws["A1"] = "RKPR RESORT - WEBSITE SOURCES REGISTER"
ws["A1"].font = title_font
ws["A1"].alignment = left_align
ws.row_dimensions[1].height = 35

# Headers
headers = [
    "Component/Section",
    "In-App Path (URL)",
    "Key Parameter",
    "Parameter Value",
    "Source File Name",
    "Source File Path",
    "Precedence Rule Applied",
    "Audit Date"
]

for col_idx, text in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col_idx, value=text)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
ws.row_dimensions[3].height = 25

# Data rows mapping parameters to sources
data = [
    # Stay GDR
    [
        "Stay - Garden Deluxe", "/stay/garden-deluxe-room",
        "Inventory / Keys", "24 Keys",
        "Garden_Deluxe_Room_Inventory.txt",
        "01_GUEST_KNOWLEDGE/Stay_Rooms/Garden_Deluxe_Room_Inventory.txt",
        "Inventory Document Rule", "2026-07-17"
    ],
    [
        "Stay - Garden Deluxe", "/stay/garden-deluxe-room",
        "Room Size", "42 sq m / 452 sq ft",
        "Garden_Deluxe_Room_Amenities.txt",
        "01_GUEST_KNOWLEDGE/Stay_Rooms/Garden_Deluxe_Room_Amenities.txt",
        "Amenities Registry Standard", "2026-07-17"
    ],
    [
        "Stay - Garden Deluxe", "/stay/garden-deluxe-room",
        "Low Season Tariff", "INR 9,500/night",
        "Garden_Deluxe_Room_Rates_and_Seasons.txt",
        "01_GUEST_KNOWLEDGE/Stay_Rooms/Garden_Deluxe_Room_Rates_and_Seasons.txt",
        "Official Rate Card Precedence", "2026-07-17"
    ],
    [
        "Stay - Garden Deluxe", "/stay/garden-deluxe-room",
        "High Season Tariff", "INR 12,500/night",
        "Garden_Deluxe_Room_Rates_and_Seasons.txt",
        "01_GUEST_KNOWLEDGE/Stay_Rooms/Garden_Deluxe_Room_Rates_and_Seasons.txt",
        "Official Rate Card Precedence", "2026-07-17"
    ],
    [
        "Stay - Garden Deluxe", "/stay/garden-deluxe-room",
        "Max Occupancy", "3 guests (2 Adults + 1 Child)",
        "Child_and_Extra_Bed_Policy.txt",
        "01_GUEST_KNOWLEDGE/Policies_Directory/Child_and_Extra_Bed_Policy.txt",
        "Unified Policy Precedence", "2026-07-17"
    ],

    # Stay VPR
    [
        "Stay - Valley View Premium", "/stay/valley-view-premium-room",
        "Inventory / Keys", "20 Keys",
        "Valley_View_Premium_Room_Inventory.txt",
        "01_GUEST_KNOWLEDGE/Stay_Rooms/Valley_View_Premium_Room_Inventory.txt",
        "Inventory Document Rule", "2026-07-17"
    ],
    [
        "Stay - Valley View Premium", "/stay/valley-view-premium-room",
        "Room Size", "48 sq m / 517 sq ft",
        "Valley_View_Premium_Room_Amenities.txt",
        "01_GUEST_KNOWLEDGE/Stay_Rooms/Valley_View_Premium_Room_Amenities.txt",
        "Amenities Registry Standard", "2026-07-17"
    ],
    [
        "Stay - Valley View Premium", "/stay/valley-view-premium-room",
        "Low Season Tariff", "INR 11,500/night",
        "Valley_View_Premium_Room_Rates_and_Seasons.txt",
        "01_GUEST_KNOWLEDGE/Stay_Rooms/Valley_View_Premium_Room_Rates_and_Seasons.txt",
        "Official Rate Card Precedence", "2026-07-17"
    ],
    [
        "Stay - Valley View Premium", "/stay/valley-view-premium-room",
        "High Season Tariff", "INR 14,500/night",
        "Valley_View_Premium_Room_Rates_and_Seasons.txt",
        "01_GUEST_KNOWLEDGE/Stay_Rooms/Valley_View_Premium_Room_Rates_and_Seasons.txt",
        "Official Rate Card Precedence", "2026-07-17"
    ],
    [
        "Stay - Valley View Premium", "/stay/valley-view-premium-room",
        "Max Occupancy", "4 guests (3 Adults or 2 Adults + 2 Children)",
        "Child_and_Extra_Bed_Policy.txt",
        "01_GUEST_KNOWLEDGE/Policies_Directory/Child_and_Extra_Bed_Policy.txt",
        "Unified Policy Precedence", "2026-07-17"
    ],

    # Stay MPS
    [
        "Stay - Mountain Panorama", "/stay/mountain-panorama-suite",
        "Inventory / Keys", "10 Keys",
        "Mountain_Panorama_Suite_Inventory.txt",
        "01_GUEST_KNOWLEDGE/Stay_Rooms/Mountain_Panorama_Suite_Inventory.txt",
        "Inventory Document Rule", "2026-07-17"
    ],
    [
        "Stay - Mountain Panorama", "/stay/mountain-panorama-suite",
        "Room Size", "68 sq m / 732 sq ft",
        "Mountain_Panorama_Suite_Amenities.txt",
        "01_GUEST_KNOWLEDGE/Stay_Rooms/Mountain_Panorama_Suite_Amenities.txt",
        "Amenities Registry Standard", "2026-07-17"
    ],
    [
        "Stay - Mountain Panorama", "/stay/mountain-panorama-suite",
        "Low Season Tariff", "INR 16,500/night",
        "Mountain_Panorama_Suite_Rates_and_Seasons.txt",
        "01_GUEST_KNOWLEDGE/Stay_Rooms/Mountain_Panorama_Suite_Rates_and_Seasons.txt",
        "Official Rate Card Precedence", "2026-07-17"
    ],
    [
        "Stay - Mountain Panorama", "/stay/mountain-panorama-suite",
        "High Season Tariff", "INR 21,000/night",
        "Mountain_Panorama_Suite_Rates_and_Seasons.txt",
        "01_GUEST_KNOWLEDGE/Stay_Rooms/Mountain_Panorama_Suite_Rates_and_Seasons.txt",
        "Official Rate Card Precedence", "2026-07-17"
    ],

    # Stay FCS
    [
        "Stay - Family Courtyard", "/stay/family-courtyard-suite",
        "Inventory / Keys", "8 Keys",
        "Family_Courtyard_Suite_Inventory.txt",
        "01_GUEST_KNOWLEDGE/Stay_Rooms/Family_Courtyard_Suite_Inventory.txt",
        "Inventory Document Rule", "2026-07-17"
    ],
    [
        "Stay - Family Courtyard", "/stay/family-courtyard-suite",
        "Room Size", "78 sq m / 840 sq ft",
        "Family_Courtyard_Suite_Amenities.txt",
        "01_GUEST_KNOWLEDGE/Stay_Rooms/Family_Courtyard_Suite_Amenities.txt",
        "Amenities Registry Standard", "2026-07-17"
    ],
    [
        "Stay - Family Courtyard", "/stay/family-courtyard-suite",
        "Low Season Tariff", "INR 19,500/night",
        "Family_Courtyard_Suite_Rates_and_Seasons.txt",
        "01_GUEST_KNOWLEDGE/Stay_Rooms/Family_Courtyard_Suite_Rates_and_Seasons.txt",
        "Official Rate Card Precedence", "2026-07-17"
    ],
    [
        "Stay - Family Courtyard", "/stay/family-courtyard-suite",
        "High Season Tariff", "INR 24,500/night",
        "Family_Courtyard_Suite_Rates_and_Seasons.txt",
        "01_GUEST_KNOWLEDGE/Stay_Rooms/Family_Courtyard_Suite_Rates_and_Seasons.txt",
        "Official Rate Card Precedence", "2026-07-17"
    ],

    # Stay HPV
    [
        "Stay - Honeymoon Villa", "/stay/honeymoon-pool-villa",
        "Inventory / Keys", "6 Keys",
        "Honeymoon_Pool_Villa_Inventory.txt",
        "01_GUEST_KNOWLEDGE/Stay_Rooms/Honeymoon_Pool_Villa_Inventory.txt",
        "Inventory Document Rule", "2026-07-17"
    ],
    [
        "Stay - Honeymoon Villa", "/stay/honeymoon-pool-villa",
        "Room Size", "92 sq m / 990 sq ft",
        "Honeymoon_Pool_Villa_Amenities.txt",
        "01_GUEST_KNOWLEDGE/Stay_Rooms/Honeymoon_Pool_Villa_Amenities.txt",
        "Amenities Registry Standard", "2026-07-17"
    ],
    [
        "Stay - Honeymoon Villa", "/stay/honeymoon-pool-villa",
        "Low Season Tariff", "INR 26,500/night",
        "Honeymoon_Pool_Villa_Rates_and_Seasons.txt",
        "01_GUEST_KNOWLEDGE/Stay_Rooms/Honeymoon_Pool_Villa_Rates_and_Seasons.txt",
        "Official Rate Card Precedence", "2026-07-17"
    ],
    [
        "Stay - Honeymoon Villa", "/stay/honeymoon-pool-villa",
        "High Season Tariff", "INR 33,500/night",
        "Honeymoon_Pool_Villa_Rates_and_Seasons.txt",
        "01_GUEST_KNOWLEDGE/Stay_Rooms/Honeymoon_Pool_Villa_Rates_and_Seasons.txt",
        "Official Rate Card Precedence", "2026-07-17"
    ],
    [
        "Stay - Honeymoon Villa", "/stay/honeymoon-pool-villa",
        "Restrictions", "Adults Only (No children/infants allowed)",
        "Smoking_and_Pet_Policy.txt",
        "01_GUEST_KNOWLEDGE/Policies_Directory/Smoking_and_Pet_Policy.txt",
        "Safety Policy Rule", "2026-07-17"
    ],

    # Stay G2PV
    [
        "Stay - Grand Two-Bedroom", "/stay/grand-two-bedroom-pool-villa",
        "Inventory / Keys", "2 Keys",
        "Grand_Two-Bedroom_Pool_Villa_Inventory.txt",
        "01_GUEST_KNOWLEDGE/Stay_Rooms/Grand_Two-Bedroom_Pool_Villa_Inventory.txt",
        "Inventory Document Rule", "2026-07-17"
    ],
    [
        "Stay - Grand Two-Bedroom", "/stay/grand-two-bedroom-pool-villa",
        "Room Size", "165 sq m / 1,776 sq ft",
        "Grand_Two-Bedroom_Pool_Villa_Amenities.txt",
        "01_GUEST_KNOWLEDGE/Stay_Rooms/Grand_Two-Bedroom_Pool_Villa_Amenities.txt",
        "Amenities Registry Standard", "2026-07-17"
    ],
    [
        "Stay - Grand Two-Bedroom", "/stay/grand-two-bedroom-pool-villa",
        "Low Season Tariff", "INR 42,000/night",
        "Grand_Two-Bedroom_Pool_Villa_Rates_and_Seasons.txt",
        "01_GUEST_KNOWLEDGE/Stay_Rooms/Grand_Two-Bedroom_Pool_Villa_Rates_and_Seasons.txt",
        "Official Rate Card Precedence", "2026-07-17"
    ],
    [
        "Stay - Grand Two-Bedroom", "/stay/grand-two-bedroom-pool-villa",
        "High Season Tariff", "INR 52,000/night",
        "Grand_Two-Bedroom_Pool_Villa_Rates_and_Seasons.txt",
        "01_GUEST_KNOWLEDGE/Stay_Rooms/Grand_Two-Bedroom_Pool_Villa_Rates_and_Seasons.txt",
        "Official Rate Card Precedence", "2026-07-17"
    ],

    # Dining
    [
        "Dining - Azure Terrace", "/dining/azure-terrace",
        "Cuisine / Service", "All-Day Dining (Indian, Asian, International)",
        "Azure_Terrace_Description.txt",
        "01_GUEST_KNOWLEDGE/Dining_Menu/Azure_Terrace_Description.txt",
        "Culinary Brochure Standard", "2026-07-17"
    ],
    [
        "Dining - Azure Terrace", "/dining/azure-terrace",
        "Breakfast Timing", "7:00 AM - 10:30 AM",
        "Azure_Terrace_Timings_and_Dresscode.txt",
        "01_GUEST_KNOWLEDGE/Dining_Menu/Azure_Terrace_Timings_and_Dresscode.txt",
        "F&B Ingestion Rule", "2026-07-17"
    ],
    [
        "Dining - Ember & Spice", "/dining/ember-spice",
        "Operation Period", "Dinner only (7:00 PM - 11:00 PM), Closed Mondays",
        "Ember_and_Spice_Timings_and_Dresscode.txt",
        "01_GUEST_KNOWLEDGE/Dining_Menu/Ember_and_Spice_Timings_and_Dresscode.txt",
        "F&B Ingestion Rule", "2026-07-17"
    ],
    [
        "Dining - Menu Catalogs", "/dining/menu",
        "Total Audited Dishes", "52 Items (Breakfast, Mains, Drinks)",
        "Various Menu Audit Logs",
        "01_GUEST_KNOWLEDGE/Dining_Menu/*_Menu_Audit.txt",
        "A La Carte Master List", "2026-07-17"
    ],

    # Spa
    [
        "Spa - Aranya Wellness", "/spa-wellness",
        "Operating Hours (Weekdays)", "9:00 AM - 9:00 PM",
        "Aranya_Wellness_Spa_Overview.txt",
        "01_GUEST_KNOWLEDGE/Spa_Wellness/Aranya_Wellness_Spa_Overview.txt",
        "Wellness Director Clearance", "2026-07-17"
    ],
    [
        "Spa - Aranya Wellness", "/spa-wellness",
        "Operating Hours (Weekends)", "8:00 AM - 10:00 PM",
        "Aranya_Wellness_Spa_Overview.txt",
        "01_GUEST_KNOWLEDGE/Spa_Wellness/Aranya_Wellness_Spa_Overview.txt",
        "Wellness Director Clearance", "2026-07-17"
    ],
    [
        "Spa - Facilities", "/spa-wellness",
        "Jacuzzi / Steam / Sauna", "Age Limit: 16+ years",
        "Aranya_Wellness_Spa_Facilities_Audit.txt",
        "01_GUEST_KNOWLEDGE/Spa_Wellness/Aranya_Wellness_Spa_Facilities_Audit.txt",
        "Spa Safety Standard", "2026-07-17"
    ],

    # Transfers
    [
        "Travel - Airport Transfer", "/transfers",
        "BLR Airport Distance", "190 km (4.5 to 5.5 hours drive)",
        "Bengaluru_Airport_Transfer_Tariffs.txt",
        "01_GUEST_KNOWLEDGE/Travel_Transfers/Bengaluru_Airport_Transfer_Tariffs.txt",
        "Logistics Route Audit", "2026-07-17"
    ],
    [
        "Travel - Sedan Tariff", "/transfers",
        "Sedan rate (BLR Airport)", "INR 8,500 one-way / INR 16,000 round-trip",
        "Bengaluru_Airport_Transfer_Tariffs.txt",
        "01_GUEST_KNOWLEDGE/Travel_Transfers/Bengaluru_Airport_Transfer_Tariffs.txt",
        "Revenue Coordinator Rule", "2026-07-17"
    ],

    # Policies
    [
        "Policies - Timings", "/directory",
        "Kids' Club Hours", "9:00 AM - 7:00 PM (Last entry 6:30 PM)",
        "Department_Timings_Directory.txt",
        "01_GUEST_KNOWLEDGE/Policies_Directory/Department_Timings_Directory.txt",
        "General Manager Guideline", "2026-07-17"
    ],
    [
        "Policies - Security Warning", "/payments",
        "OTP / PIN / CVV safeguard", "Never request OTP, CVV, Card PIN, UPI PIN",
        "Billing_and_Payments_Policy.txt",
        "01_GUEST_KNOWLEDGE/Policies_Directory/Billing_and_Payments_Policy.txt",
        "Legal Counsel Mandated Disclaimer", "2026-07-17"
    ]
]

# Write data
row_idx = 4
for row in data:
    for col_idx, value in enumerate(row, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = cell_font
        cell.border = thin_border
        if col_idx in [1, 2, 5, 6, 7]:
            cell.alignment = left_align
        elif col_idx in [3, 4]:
            cell.alignment = left_align
            # Highlight key parameter slightly
            if col_idx == 3:
                cell.font = bold_font
        else:
            cell.alignment = center_align
    row_idx += 1

from openpyxl.utils import get_column_letter

# Auto-adjust column widths
for col in ws.columns:
    max_len = 0
    col_idx = col[0].column
    col_letter = get_column_letter(col_idx)
    for cell in col:
        if cell.row == 1:
            continue
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

# Save workbook
wb.save(output_path)
print(f"Spreadsheet generated successfully at: {output_path}")
